#=CONCAT("https://sweetregina.com/wp-content/uploads/2021/8/",SUBSTITUTE(D2," ","-"),".jpg")
from openpyxl import load_workbook
import os

def find(name):
    for root, dirs, files in os.walk("sweetReginaPhotos"):
        if name in files:
            return os.path.join(root, name)
wb = load_workbook('prods-sweet.xlsx')
ws = wb.active
for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
    for cell in row:
        print(cell.row)
        if find(cell.value+'.png') is not None:
            ws['AD'+str(cell.row)] = 'https://sweetregina.com/wp-content/uploads/2021/08/'+cell.value.lstrip().replace(" ","-").replace("'","").replace("ñ","n").replace("á","a").replace("é","e").replace("í","í").replace("ó","o").replace("ú","u")+'.jpg'
        else:
            ws['AD'+str(cell.row)] = None
wb.save('prods-new.xlsx')
wb.close()